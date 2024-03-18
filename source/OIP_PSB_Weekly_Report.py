import os
from datetime import datetime, date
import shutil
from pyspark.sql import SparkSession
from pyspark.sql.types import StructType, StructField, StringType
from pyspark.sql.functions import lit, current_date, to_date
import pyodbc
import pandas as pd
from SGTAMProdTask import SGTAMProd
import logging
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side
import openpyxl

################################################################################################################
# Declare custom exception for missing F file check
class ExceptionMissingFFile(Exception):
    pass
################################################################################################################


# Get today's date
today_date = datetime.today().date()

# Convert to the specified format
formatted_today_date = today_date.strftime('%Y-%m-%d')

# map directly to programlog download folder
directory_path = 'J:\\'

# to store the daily folders created in the "downloads"
folders_downloads = []

# to store the latest f file name
latest_F_file = None

# Set up logging
log_filename = f"D:/SGTAM_DP/Working Project/Weekly OIP PSB Program Title Report/source/log/WeeklyPSBProgramTitleReport_{datetime.now().strftime('%Y-%m-%d %H-%M-%S')}.txt"
logging.basicConfig(filename=log_filename, level=logging.INFO)
s = SGTAMProd()


try:
    # -----------------------------------------------------------------------------------------

    print('Start of code block for getting the list of files names in the daily folder in download folder')
    logging.info('Start of code block for getting the list of files names in the daily folder in download folder')
    # Get a list of all items (files and folders) in the directory
    items_in_directory = os.listdir(directory_path)

    # Filter out the folders that start with "2024"
    # folders_starting_with_2023 = [folder for folder in items_in_directory if os.path.isdir(os.path.join(directory_path, folder)) and folder.startswith("2023")]
    folders_starting_with_2024 = [folder for folder in items_in_directory if os.path.isdir(os.path.join(directory_path, folder)) and folder.startswith("2024")]

    # Display the list of folders starting with "2024"
    print("Folders starting with '2024':")
    for folder in folders_starting_with_2024:
        folders_downloads.append(folder)
        print(folder)

    # Convert the date strings to datetime objects
    date_objects = [datetime.strptime(date_str, "%Y-%m-%d") for date_str in folders_downloads]

    # Find the maximum date
    max_date = str(max(date_objects).strftime("%Y-%m-%d"))

    # Print the maximum date without the time
    print("Max date found in downloads folder:", max_date)
    logging.info(f"Max date found in downloads folder: {max_date}")

    
    print('Getting a list of files in the latest daily folder')
    logging.info('Getting a list of files in the latest daily folder')
    # Get a list of files in the directory
    files_in_directory = [file for file in os.listdir(f"{directory_path}\\{max_date}") if os.path.isfile(os.path.join(f"{directory_path}\\{max_date}", file)) and file.startswith("F24")]

    print(files_in_directory)
    logging.info(files_in_directory)


    print('Checking if the F file is available in the latest daily folder.\nIf it is not available it will raise exception and send out warning email.')
    logging.info('Checking if the F file is available in the latest daily folder.\nIf it is not available it will raise exception and send out warning email.')

    if not files_in_directory:
        print("No F prelog file(s) found.")
        logging.info("No F prelog file(s) found.")
        print(f"Raising custom exception to ExceptionMissingFFile.")
        logging.info(f"Raising custom exception to ExceptionMissingFFile.")
        raise ExceptionMissingFFile('No F file(s) available.')
        

    else:
        print('F file is available, continuing the process.')
        logging.info('F file is available, continuing the process.')

        # Define a function to extract the date and version from the file name
        def extract_date_and_version(file_name):
            date_str = file_name[1:7]  # Extract the date part (e.g., '230719') from the file name
            version_str = file_name[7:-4]  # Extract the version part (e.g., 'A') from the file name
            return datetime.strptime(date_str, "%d%m%y"), version_str

        print('Getting a list of F files and sort them from latest to oldest version')
        logging.info('Getting a list of F files and sort them from latest to oldest version')
        # Sort the files based on date and version
        sorted_files = sorted(files_in_directory, key=extract_date_and_version, reverse=True)
        print(f'Available F files today: {sorted_files}')
        logging.info(f'Available F files today: {sorted_files}')



        # Get the latest file
        latest_F_file = sorted_files[0]
        print("Latest F prelog file:", latest_F_file)
        logging.info(f'Latest F prelog file: {latest_F_file}')
        print(f"Copy the latest F file {latest_F_file} to local drive")
        logging.info(f"Copy the latest F file {latest_F_file} to local drive")
        f_file_source = os.path.join(directory_path, max_date, latest_F_file)
        f_file_copy_destination = f'D:/SGTAM_DP/Working Project/Weekly OIP PSB Program Title Report/prelogFiles'
        shutil.copy2(f_file_source, f_file_copy_destination)
        print(f'Copied {latest_F_file} to {f_file_copy_destination}')
        logging.info(f'Copied {latest_F_file} to {f_file_copy_destination}')



        # Check if the same filename and import date already exist in the table, if exist then will not insert/update the table
        print('Creating SQL connection to SGTAMProdOIP')
        logging.info('Creating SQL connection to SGTAMProdOIP')
        # SQL Connection Infos
        server = 'xxx'
        database = 'xxx'
        username = 'xxx'
        password = 'xxx'
        driver = '{ODBC Driver 17 for SQL Server}'
        # Set up SQL connection
        cnxn = pyodbc.connect(f"DRIVER={driver};SERVER={server};DATABASE={database};UID={username};PWD={password}")
        cursor = cnxn.cursor()

        # Check for existing records
        query_check = """
        SELECT COUNT(*) FROM tOIPPreLog3
        WHERE file_name = ? AND import_date = ?
        """

        print('Executing query to check if there are any records that have the same file_name and import_date as today import')
        logging.info('Executing query to check if there are any records that have the same file_name and import_date as today import')
        cursor.execute(query_check, latest_F_file[0:8], formatted_today_date)
        count = cursor.fetchone()[0]

        if count > 0:
            print(f"{count} records found in tOIPPreLog3 with file_name={latest_F_file[0:8]} and import_date={formatted_today_date}. Deleting records with these file_name and import_date.")
            logging.info(f"{count} records found in tOIPPreLog3 with file_name={latest_F_file[0:8]} and import_date={formatted_today_date}. Deleting records with these file_name and import_date.")

            # Delete the records
            query_delete = """
            DELETE FROM tOIPPreLog3
            WHERE file_name = ? AND import_date = ?
            """

            print('Removing those records from SGTAMProdOIP')
            logging.info('Removing those records from SGTAMProdOIP')
            cursor.execute(query_delete, latest_F_file[0:8], formatted_today_date)
            cnxn.commit()
            print("Records deleted")
            logging.info('Records deleted')
        else:
            print(f"No records found with file_name={latest_F_file[0:8]} and import_date={formatted_today_date}.\n Proceeding to insert the data.")
            logging.info(f"No records found with file_name={latest_F_file[0:8]} and import_date={formatted_today_date}.\n Proceeding to insert the data.")

        # Close the cursor and connection
        cursor.close()
        cnxn.close()
        print('SQL connection closed.')
        logging.info('SQL connection closed.')



        # Create a SparkSession
        print('Create new Sparks session')
        logging.info('Create new Sparks session')
        spark = SparkSession.builder \
            .appName('Read Text File') \
            .config("spark.driver.extraClassPath", "D:\\SGTAM_DP\\Working Project\\Weekly OIP PSB Program Title Report\\source\\mssql-jdbc-12.4.0.jre11.jar") \
            .getOrCreate()
        
        
        # create dataframe from the file
        csv_file_path = os.path.join(f_file_copy_destination, latest_F_file)

        # Define the schema / columns, datatype and if its nullable, must match that of the SQL DATABASE
        schema = StructType([
            StructField("CHANNEL_ID", StringType(), True),
            StructField("CHANNEL_BELT", StringType(), True),
            StructField("TX_DATE", StringType(), True),
            StructField("START_TIME", StringType(), True),
            StructField("SLOT_DURATION", StringType(), True),
            StructField("SLOT_NAME", StringType(), True),
            StructField("MAIN_TITLE", StringType(), True),
            StructField("EPISODE_TITLE", StringType(), True),
            StructField("GENRE", StringType(), True),
            StructField("SUB_GENRE", StringType(), True),
            StructField("REPEAT", StringType(), True),
            StructField("LIVE_FLAG", StringType(), True),
            StructField("SUBTITLE_LANGUAGE", StringType(), True),
            StructField("PROGRAMME_ID", StringType(), True),
            StructField("COUNTRY_DESC", StringType(), True),
            StructField("LANGUAGE_DESC", StringType(), True),
            StructField("PRODUCTION_TYPE_DESC", StringType(), True),
            StructField("LOADING_FACTOR", StringType(), True),
            StructField("SPONSORED_FLAG", StringType(), True),
            StructField("LIVE_PROGRAMME_ID", StringType(), True),
            StructField("REFERENCE_ID", StringType(), True),
            StructField("MASTER_REFERENCE_KEY", StringType(), True),
            StructField("PRODUCTION_SUB_TYPE", StringType(), True),
            StructField("PSB_SURVEY", StringType(), True),
            StructField("ALTERNATE_LANGUAGE", StringType(), True)
        ])


        print(f'Creating PySpark dataframe from the F file {latest_F_file}')
        logging.info(f'Creating PySpark dataframe from the F file {latest_F_file}')
        # header = false , mean dont treat first line of the file as the headers
        df = spark.read.format("csv") \
            .option("delimiter", "\t") \
            .option("header", "false") \
            .schema(schema) \
            .load(csv_file_path)


        # Add columns "ref_date" and "import_date" with default values to the DataFrame
        df = df.withColumn("file_name", lit(latest_F_file[0:8]).cast(StringType())) \
               .withColumn("import_date", current_date())

        # Reorder the columns to have the new columns at the first and second positions
        column_order = ["file_name", "import_date"] + df.columns[:-2]
        df = df.select(*column_order)

        # Converting the TX_DATE date string to a date type 
        df = df.withColumn("TX_DATE", to_date(df["TX_DATE"], 'yyyyMMdd'))
        
        print('PySpark dataframe created.')
        logging.info('PySpark dataframe created.')

        # check how many records in dataframe
        row_count = df.count()
        print(f'\nDataframe has {row_count} rows, please cross check with the data file')
        logging.info(f'\nDataframe has {row_count} rows, please cross check with the data file')

        #-------------------------------------------------------------------------------------------------------
        count = 0
        with open(csv_file_path, 'r', encoding='utf-8') as file:
            for line in file:
                if line.strip():  # Check if the line contains any non-whitespace characters
                    count += 1

        print(f"{latest_F_file} has {count} rows of non-empty lines\n")
        logging.info(f"{latest_F_file} has {count} rows of non-empty lines\n")
         #-------------------------------------------------------------------------------------------------------


        print(f'Importing the dataframe with file_name:{latest_F_file} and import_date:{formatted_today_date} into tOIPPreLog3')
        logging.info(f'Importing the dataframe with file_name:{latest_F_file} and import_date:{formatted_today_date} into tOIPPreLog3')
        df.write \
          .format("jdbc") \
          .option("url", "jdbc:sqlserver:xxx:1433;databaseName=xxx;trustServerCertificate=true") \
          .option("dbtable", "xxx") \
          .option("user", "xxx") \
          .option("password", "xxx") \
          .mode("append") \
          .save()

        print('Imported completed')
        logging.info('Imported completed')

        # Closing the SparkSession
        spark.stop()
        print('PySpark connection stopped')
        logging.info('PySpark connection stopped')


        ##--------------------------------------------------------------------------------------------------------------##
        ## Get results from stored prod and export to excel file
        ##--------------------------------------------------------------------------------------------------------------##
        
        # Create a Spark session
        spark = SparkSession.builder \
            .appName('SQL Server Connection') \
            .config("spark.driver.extraClassPath", "D:\\SGTAM_DP\\Working Project\\Weekly OIP PSB Program Title Report\\source\\mssql-jdbc-12.4.0.jre11.jar") \
            .getOrCreate()
        print('Created spark connection again.')
        logging.info('Created spark connection again.')

        
        
        # Define queries to call the stored procedures
        query_main_report = f"EXEC SP_OIP_PSB_Weekly_Report_Main '{latest_F_file[0:8]}', '{formatted_today_date}'"
        query_secondary_report = f"EXEC SP_OIP_PSB_Weekly_Report_Secondary '{latest_F_file[0:8]}', '{formatted_today_date}'"

        # Create a connection using pyodbc
        print('Creating pandas dataframes from the SQL results for main and secondary reports')
        logging.info('Creating pandas dataframes from the SQL results for main and secondary reports')
        with pyodbc.connect(f"DRIVER={driver};SERVER={server};DATABASE={database};UID={username};PWD={password}") as cnxn:
            pandas_df_main = pd.read_sql(query_main_report, cnxn)
            pandas_df_secondary = pd.read_sql(query_secondary_report, cnxn)
            # Replace all NaN or empty values with 'NULL' in the DataFrames
            pandas_df_main.fillna('NULL', inplace=True)
            pandas_df_secondary.fillna('NULL', inplace=True)

        # Convert the column to datetime if it's not already
        pandas_df_main['Broadcast Date'] = pd.to_datetime(pandas_df_main['Broadcast Date'])
        pandas_df_main['Start Date'] = pd.to_datetime(pandas_df_main['Start Date'])
        pandas_df_main['End Date'] = pd.to_datetime(pandas_df_main['End Date'])
        pandas_df_secondary['import_date'] = pd.to_datetime(pandas_df_secondary['import_date'])
        pandas_df_secondary['TX_DATE'] = pd.to_datetime(pandas_df_secondary['TX_DATE'])

        # # Convert the column from datetime to date 
        # pandas_df_main['Broadcast Date'] = pandas_df_main['Broadcast Date'].dt.date
        # pandas_df_main['Start Date'] = pandas_df_main['Start Date'].dt.date
        # pandas_df_main['End Date'] = pandas_df_main['End Date'].dt.date
        # pandas_df_secondary['import_date'] = pandas_df_secondary['import_date'].dt.date
        # pandas_df_secondary['TX_DATE'] = pandas_df_secondary['TX_DATE'].dt.date

        # Change the date format to dd/mm/yyyy
        pandas_df_main['Broadcast Date'] = pandas_df_main['Broadcast Date'].dt.strftime('%d/%m/%Y')
        pandas_df_main['Start Date'] = pandas_df_main['Start Date'].dt.strftime('%d/%m/%Y')
        pandas_df_main['End Date'] = pandas_df_main['End Date'].dt.strftime('%d/%m/%Y')
        pandas_df_secondary['import_date'] = pandas_df_secondary['import_date'].dt.strftime('%d/%m/%Y')
        pandas_df_secondary['TX_DATE'] = pandas_df_secondary['TX_DATE'].dt.strftime('%d/%m/%Y')


        # Convert these columns to Int64 nullable type to prevent type error in excel
        pandas_df_main['Program ID'] = pandas_df_main['Program ID'].astype('Int64')
        pandas_df_main['Content ID'] = pandas_df_main['Content ID'].astype('Int64') 
        pandas_df_secondary['START_TIME'] = pandas_df_secondary['START_TIME'].astype('Int64')
        pandas_df_secondary['SLOT_DURATION'] = pandas_df_secondary['SLOT_DURATION'].astype('Int64')
        # pandas_df_secondary['PROGRAMME_ID'] = pandas_df_secondary['PROGRAMME_ID'].astype('Int64') ##test
        pandas_df_secondary['PROGRAMME_ID'] = pandas_df_secondary['PROGRAMME_ID'].apply(pd.to_numeric, errors='ignore') ##test #due to Asian Games
        pandas_df_secondary['LOADING_FACTOR'] = pandas_df_secondary['LOADING_FACTOR'].astype('Int64')

        # # Convert Pandas DataFrames to PySpark DataFrames if needed
        # df_main = spark.createDataFrame(pandas_df_main)
        # df_secondary = spark.createDataFrame(pandas_df_secondary)
        # print('Convereted the pandas dataframe PySpark dataframes')
        # logging.info('Convereted the pandas dataframe PySpark dataframes')

        # Create an Excel writer with Pandas
        print('Creating the excel report file.')
        logging.info('Creating the excel report file.')
        writer = pd.ExcelWriter("PSB Program Title Report.xlsx", engine='openpyxl', date_format='MM/DD/YYYY')
                
        # Write the Pandas DataFrames to two different sheets in the Excel file
        pandas_df_main.to_excel(writer, sheet_name='Main Report', index=False)
        print('Main report written to excel file')
        logging.info('Main report written to excel file')
        
        pandas_df_secondary.to_excel(writer, sheet_name=f'{latest_F_file[0:8]}', index=False)
        print('Secondary report written to excel file')
        logging.info('Secondary report written to excel file')

        # Save the Excel file
        # writer.save()
        writer.close()
        print('Excel report file saved')
        logging.info('Excel report file saved')

        #-----------------------------------------------------------------------------------------------------------------------------------#
        # Access excel file and apply format to a few date columns
        # Define the path and load the workbook
        excel_file_path = 'D:/SGTAM_DP/Working Project/Weekly OIP PSB Program Title Report/source/PSB Program Title Report.xlsx'
        workbook = openpyxl.load_workbook(excel_file_path)

        # Define the date columns for each sheet
        sheet_columns_map = {
            "Main Report": ['Broadcast Date', 'Start Date', 'End Date'],
            f'{latest_F_file[0:8]}': ['import_date', 'TX_DATE']
        }
        
        # Define the date style
        date_style = NamedStyle(name='date_style', number_format='MM/DD/YYYY')
        
        # Function to process each sheet and format date columns
        def format_date_columns(sheet, columns):
            for col_name in columns:
                col_index = None
                for idx, cell in enumerate(sheet[1], 1):
                    if cell.value == col_name:
                        col_index = idx
                        break
                    
                # If column is found, process it
                if col_index:
                    # Set column width
                    sheet.column_dimensions[openpyxl.utils.get_column_letter(col_index)].width = 15
                    
                    # Process each cell in the column
                    for row in sheet.iter_rows(min_row=2, min_col=col_index, max_col=col_index):
                        for cell in row:
                            cell_value = cell.value
        
                            # Convert string to date format
                            if isinstance(cell_value, str):
                                try:
                                    date_value = pd.to_datetime(cell_value, dayfirst=True).date()
                                    cell.value = date_value
                                    cell.style = date_style
                                except ValueError:
                                    # Silently pass for non-date strings
                                    pass
                                
                            # Convert pd.Timestamp to date format
                            elif isinstance(cell_value, pd.Timestamp):
                                date_value = cell_value.date()
                                cell.value = date_value
                                cell.style = date_style
        
        # Apply formatting to sheets
        for sheet_name, columns in sheet_columns_map.items():
            format_date_columns(workbook[sheet_name], columns)
        
        
        
        # List of strings to populate 'Exclusive Titles' sheet
        # (You can fill this list with your actual values)
        titles = ["What on Earth S2", "Let Me Tell You A Story", "Streets Made For Talking : Telok Ayer", "Oh Butterfly!", "Measuring Meritocracy", "Space Farmers", "MasterChef Singapore Season 4", "The Great Migration: New Eden", "Pesuvom Sr 2","Lights. Camera. Action On Caldecott","The Roots Of Our Garden","Untold Legends","Premium Rush: Inside Air Cargo Singapore - Part 1","Premium Rush: Inside Air Cargo Singapore - Part 2","ROOTS: A Greening Journey"]  # Replace ... with your list of strings

        # Create a new sheet 'Exclusive Titles' 
        if "Exclusive Titles" in workbook.sheetnames:
            ws_exclusive_titles = workbook["Exclusive Titles"]
        else:
            ws_exclusive_titles = workbook.create_sheet("Exclusive Titles")

        # Add the header to column A
        ws_exclusive_titles["A1"] = "Main Title"
        bold_font = Font(bold=True)
        ws_exclusive_titles["A1"].font = bold_font

        # Define a full border (all sides)
        all_border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
        ws_exclusive_titles["A1"].border = all_border   

        # Populate the cells in column A starting from the second row
        for idx, title in enumerate(titles, start=2):
            ws_exclusive_titles[f"A{idx}"] = title


        # Save the changes
        workbook.save(excel_file_path)

        #-----------------------------------------------------------------------------------------------------------------------------------#


        # Stop the Spark session
        spark.stop()
        print('PySpark connection stopped')
        logging.info('PySpark connection stopped')

        # Send missing F file email
        logging.info("Preparing to send successful email")
        print("Preparing to send successful email")
        email_body = f"<p>Hi all,</p><p>Kindly find the weekly PSB Program Title Report, it was generated based on the latest prelog file {latest_F_file} received today.</p><br><p>*This is an auto-generated email, kindly reply to SGTAMDPTeam@gfk.com instead.*</p>"
        email_kwargs = {
            'sender':'xxx',
            'to':'xxx',
            'subject':f'PSB Program Title Report - {formatted_today_date}',
            'body':email_body,
            'is_html':True,
            'filename':"D:/SGTAM_DP/Working Project/Weekly OIP PSB Program Title Report/source/PSB Program Title Report.xlsx"
        }
        s.send_email(**email_kwargs)
        logging.info("Report sent")
        print("Report sent")

# To send WARNING email for missing F file
except ExceptionMissingFFile as e:
    print(e)
    logging.info(e)
    # Send missing F file email
    logging.info("Preparing to send warning email")
    print("Preparing to send warning email")
    email_body = f"<p>It seems there is no prelog file available at the moment, please restart the process once the file is available</p><p>{e}</p><p>Please check log at {log_filename}</p><p>*This is an auto generated email, do not reply to this email.</p>"
    email_kwargs = {
        'sender':'xxx',
        'to':'xxx',
        'subject':f'[WARNING] OIP PSB Weekly Report - {formatted_today_date}',
        'body':email_body,
        'is_html':True,
        'filename':log_filename
    }
    s.send_email(**email_kwargs)
    logging.info("Warning email sent")
    print("Warning email sent")

except Exception as e:
    print(f'There is an error:\n{e}')
    logging.info(f'There is an error:\n{e}')
    # Send error email
    logging.info("Preparing to send error email")
    print("Preparing to send error email")
    email_body = f"<p>There is an error:</p><p>{e}</p><p>Please check log at {log_filename}</p><p>*This is an auto generated email, do not reply to this email.</p>"
    email_kwargs = {
        'sender':'xxx',
        'to':'xxx',
        'subject':f'[ERROR] OIP PSB Weekly Report - {formatted_today_date}',
        'body':email_body,
        'is_html':True,
        'filename':log_filename
    }
    s.send_email(**email_kwargs)
    logging.info("Error email sent")
    print("Error email sent")

finally:
    print('This is the finally clause.')
    logging.info('This is the finally clause.')